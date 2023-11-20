# utils/excel_reader.py
# utilities/excel_reader.py
import openpyxl


class ExcelReader:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook[sheet_name]

    def read_data(self):
        column_names = [cell.value for cell in self.sheet[1]]
        data = []

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(column_names, row)))

        return data

    def write_data(self, row, column, value):
        self.sheet.cell(row=row, column=column).value = value
        self.workbook.save(self.file_path)
